from __future__ import annotations

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from bson.objectid import ObjectId
from datetime import datetime, timedelta
from typing import Dict, Any, Tuple, Optional, List
from io import BytesIO
import os
import threading
import uuid

import requests
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except Exception:
    openpyxl = None

from db import db

executive_auditing_bp = Blueprint("executive_auditing", __name__, url_prefix="/executive/auditing")

users_col = db["users"]
audits_col = db["executive_audits"]
payments_col = db["payments"]
customers_col = db["customers"]
archived_customers_col = db["Archived_customers"]
packages_col = db["packages"]
card_closures_col = db["card_closures"]

LOGO_URL = "https://imagedelivery.net/h9fmMoa1o2c2P55TcWJGOg/10283f28-b9a0-4100-b254-19a854386800/public"
COMPANY_NAME = "SMARTLIVING"

JOB_LOCK = threading.Lock()
JOBS: Dict[str, Dict[str, Any]] = {}
EXPORT_DIR = os.path.join(os.getcwd(), "uploads", "audit_exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _safe_float(val, default: float = 0.0) -> float:
    try:
        return float(val)
    except Exception:
        return default


def _safe_int(val, default: int = 0) -> int:
    try:
        return int(val)
    except Exception:
        return default


def _current_exec_session() -> Tuple[Optional[str], Optional[str]]:
    if session.get("executive_id"):
        return "executive_id", session["executive_id"]
    if session.get("admin_id"):
        return "admin_id", session["admin_id"]
    return None, None


def _ensure_exec_or_redirect():
    _, uid = _current_exec_session()
    if not uid:
        return redirect(url_for("login.login"))
    try:
        user = users_col.find_one({"_id": ObjectId(uid)})
    except Exception:
        user = users_col.find_one({"_id": uid})
    if not user:
        return redirect(url_for("login.login"))
    role = (user.get("role") or "").lower()
    if role not in ("executive", "admin"):
        return redirect(url_for("login.login"))
    return str(user["_id"]), user


def _subject_label(user: Dict[str, Any]) -> str:
    return user.get("name") or user.get("username") or "User"


def _build_match(args) -> Dict[str, Any]:
    match: Dict[str, Any] = {}
    role = (args.get("role") or "").strip().lower()
    status = (args.get("status") or "").strip().title()

    if role in ("manager", "admin", "agent"):
        match["subject_role"] = role
    if status in ("Open", "Settled", "Partial"):
        match["status"] = status
    return match


def _parse_range(range_key: str, start_str: str, end_str: str) -> Tuple[Optional[datetime], Optional[datetime], str]:
    range_key = (range_key or "all_time").lower()
    if range_key == "last_6_months":
        end = datetime.utcnow()
        start = end - timedelta(days=182)
        label = f"Last 6 months ({start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')})"
        return start, end, label
    if range_key == "custom":
        try:
            start = datetime.strptime(start_str, "%Y-%m-%d")
        except Exception:
            start = None
        try:
            end = datetime.strptime(end_str, "%Y-%m-%d")
        except Exception:
            end = None
        if start and end:
            label = f"Custom ({start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')})"
        else:
            label = "Custom"
        return start, end, label
    return None, None, "All time"


def _safe_date_str(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value[:10]
    return ""


def _agent_doc(agent_id: str) -> Optional[Dict[str, Any]]:
    if not agent_id:
        return None
    try:
        return users_col.find_one({"_id": ObjectId(agent_id)})
    except Exception:
        return users_col.find_one({"_id": agent_id})


def _manager_doc(manager_id: Any) -> Optional[Dict[str, Any]]:
    if not manager_id:
        return None
    try:
        return users_col.find_one({"_id": ObjectId(manager_id)})
    except Exception:
        return users_col.find_one({"_id": manager_id})


def _id_variants(value: Any) -> List[Any]:
    if not value:
        return []
    values: List[Any] = [value, str(value)]
    try:
        oid = value if isinstance(value, ObjectId) else ObjectId(str(value))
        values.append(oid)
    except Exception:
        pass
    deduped: List[Any] = []
    for item in values:
        if item not in deduped:
            deduped.append(item)
    return deduped


def _customer_id_values(customer_ids: List[ObjectId]) -> List[Any]:
    return customer_ids + [str(cid) for cid in customer_ids]


def _chunks(items: List[Any], size: int = 80):
    for index in range(0, len(items), size):
        yield items[index:index + size]


def _card_key(customer_id: Any, product_index: Any) -> Tuple[str, int]:
    return (str(customer_id), _safe_int(product_index, -1))


def _fmt_money(value: Any) -> str:
    return f"{_safe_float(value):,.2f}"


def _build_product_payment_map(customer_ids: List[ObjectId]) -> Dict[Tuple[str, int], float]:
    if not customer_ids:
        return {}
    payment_map: Dict[Tuple[str, int], float] = {}
    for batch in _chunks(customer_ids):
        cursor = payments_col.find(
            {
                "customer_id": {"$in": _customer_id_values(batch)},
                "product_index": {"$ne": None},
                "payment_type": {"$nin": ["SUSU"]},
            },
            {"customer_id": 1, "product_index": 1, "payment_type": 1, "amount": 1},
        )
        for payment in cursor:
            key = _card_key(payment.get("customer_id"), payment.get("product_index"))
            if key[1] < 0:
                continue
            amount = _safe_float(payment.get("amount"), 0.0)
            if payment.get("payment_type") == "WITHDRAWAL":
                amount *= -1
            payment_map[key] = payment_map.get(key, 0.0) + amount
    return payment_map


def _build_package_status_map(customer_ids: List[ObjectId]) -> Dict[Tuple[str, int], Dict[str, Any]]:
    if not customer_ids:
        return {}
    package_map: Dict[Tuple[str, int], Dict[str, Any]] = {}
    cursor = packages_col.find(
        {"customer_id": {"$in": _customer_id_values(customer_ids)}, "status": {"$ne": "cancelled"}},
        {"customer_id": 1, "product_index": 1, "status": 1, "created_at": 1},
    ).sort("created_at", -1)
    for package in cursor:
        key = _card_key(package.get("customer_id"), package.get("product_index"))
        if key[1] < 0 or key in package_map:
            continue
        package_map[key] = {
            "status": str(package.get("status") or "").strip().lower(),
            "created_at": package.get("created_at"),
        }
    return package_map


def _build_closed_index_set(customer_ids: List[ObjectId]) -> set[Tuple[str, int]]:
    if not customer_ids:
        return set()
    closed: set[Tuple[str, int]] = set()
    for doc in card_closures_col.find(
        {"customer_id": {"$in": _customer_id_values(customer_ids)}, "action": "close_card"},
        {"customer_id": 1, "payload.selected_product_index": 1},
    ):
        key = _card_key(doc.get("customer_id"), (doc.get("payload") or {}).get("selected_product_index"))
        if key[1] >= 0:
            closed.add(key)
    return closed


def _lookup_users_by_ids(ids: set[str]) -> Dict[str, Dict[str, Any]]:
    oid_list = [ObjectId(x) for x in ids if ObjectId.is_valid(str(x))]
    if not oid_list:
        return {}
    out: Dict[str, Dict[str, Any]] = {}
    for user in users_col.find({"_id": {"$in": oid_list}}, {"name": 1, "username": 1, "branch": 1, "branch_name": 1, "manager_id": 1}):
        out[str(user["_id"])] = user
    return out


def _build_customer_card_tabs(agent_id: Optional[str] = None, limit: Optional[int] = None) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Dict[str, Any]]]:
    customer_query: Dict[str, Any] = {"purchases.0": {"$exists": True}}
    if agent_id:
        customer_query["agent_id"] = {"$in": _agent_id_variants(agent_id)}

    cursor = (
        customers_col.find(
            customer_query,
            {
                "_id": 1,
                "name": 1,
                "phone_number": 1,
                "location": 1,
                "status": 1,
                "branch": 1,
                "agent_id": 1,
                "manager_id": 1,
                "purchases": 1,
                "date_registered": 1,
            },
        )
        .sort([("_id", -1)])
    )
    if limit:
        cursor = cursor.limit(limit)
    customers = list(cursor)
    customer_ids = [c["_id"] for c in customers]
    payment_map = _build_product_payment_map(customer_ids)
    package_map = _build_package_status_map(customer_ids)
    closed_set = _build_closed_index_set(customer_ids)

    user_ids: set[str] = set()
    for customer in customers:
        if customer.get("agent_id"):
            user_ids.add(str(customer.get("agent_id")))
        if customer.get("manager_id"):
            user_ids.add(str(customer.get("manager_id")))
    users_by_id = _lookup_users_by_ids(user_ids)

    tabs: Dict[str, List[Dict[str, Any]]] = {"ongoing": [], "closed": [], "completed": []}
    completed_statuses = {
        "completed",
        "submitted_for_packaging",
        "packaged",
        "packaging",
        "delivering",
        "delivered",
    }

    for customer in customers:
        cid = customer["_id"]
        agent_doc = users_by_id.get(str(customer.get("agent_id"))) or {}
        manager_doc = users_by_id.get(str(customer.get("manager_id"))) or {}
        branch = (
            manager_doc.get("branch_name")
            or manager_doc.get("branch")
            or customer.get("branch")
            or agent_doc.get("branch")
            or ""
        )
        for index, purchase in enumerate(customer.get("purchases") or []):
            product = (purchase or {}).get("product") or {}
            product_name = product.get("name") or "Unknown Product"
            product_status = str(product.get("status") or purchase.get("status") or "active").strip().lower()
            package_info = package_map.get(_card_key(cid, index)) or {}
            package_status = str(package_info.get("status") or "").strip().lower()
            total = _safe_float(product.get("total"), 0.0)
            paid = round(payment_map.get(_card_key(cid, index), 0.0), 2)
            left = max(0.0, round(total - paid, 2))
            is_closed = product_status == "closed" or purchase.get("status") == "closed" or _card_key(cid, index) in closed_set
            is_completed = (not is_closed) and (
                left <= 0
                or product_status in completed_statuses
                or (package_status and package_status != "cancelled")
            )
            bucket = "closed" if is_closed else ("completed" if is_completed else "ongoing")

            tabs[bucket].append({
                "customer_id": str(cid),
                "customer_name": customer.get("name") or "Unknown Customer",
                "phone": customer.get("phone_number") or "",
                "location": customer.get("location") or "",
                "customer_status": customer.get("status") or "",
                "product_index": index,
                "product_name": product_name,
                "product_status": product_status or "active",
                "package_status": package_status,
                "branch": branch,
                "agent_name": agent_doc.get("name") or agent_doc.get("username") or "",
                "manager_name": manager_doc.get("name") or manager_doc.get("username") or "",
                "total": total,
                "paid": paid,
                "left": left,
                "purchase_date": purchase.get("purchase_date") or "",
            })

    def _sort_key(row: Dict[str, Any]):
        return (row.get("left", 0.0), row.get("customer_name", ""))

    tabs["ongoing"].sort(key=_sort_key, reverse=True)
    tabs["closed"].sort(key=lambda r: (r.get("purchase_date") or "", r.get("customer_name") or ""), reverse=True)
    tabs["completed"].sort(key=lambda r: (r.get("left", 0.0), r.get("customer_name") or ""))

    tab_stats: Dict[str, Dict[str, Any]] = {}
    for key, rows in tabs.items():
        tab_stats[key] = {
            "count": len(rows),
            "total": sum(_safe_float(r.get("total")) for r in rows),
            "paid": sum(_safe_float(r.get("paid")) for r in rows),
            "left": sum(_safe_float(r.get("left")) for r in rows),
        }
    return tabs, tab_stats


def _payments_stats(agent_id: str, start: Optional[datetime], end: Optional[datetime]) -> Dict[str, Any]:
    match: Dict[str, Any] = {
        "agent_id": agent_id,
        "payment_type": {"$ne": "WITHDRAWAL"}
    }

    pipeline: List[Dict[str, Any]] = [
        {"$match": match},
        {"$addFields": {
            "pay_date": {
                "$ifNull": [
                    "$timestamp",
                    {"$dateFromString": {
                        "dateString": "$date",
                        "format": "%Y-%m-%d",
                        "onError": None,
                        "onNull": None
                    }}
                ]
            }
        }},
    ]
    if start and end:
        pipeline.append({"$match": {"pay_date": {"$gte": start, "$lte": end}}})

    pipeline.append({
        "$group": {
            "_id": None,
            "total_amount": {"$sum": {"$ifNull": ["$amount", 0]}},
            "count": {"$sum": 1},
            "min_date": {"$min": "$pay_date"},
            "max_date": {"$max": "$pay_date"},
        }
    })

    agg = list(payments_col.aggregate(pipeline))
    stats = agg[0] if agg else {"total_amount": 0, "count": 0, "min_date": None, "max_date": None}
    return {
        "total_amount": float(stats.get("total_amount", 0) or 0),
        "count": int(stats.get("count", 0) or 0),
        "min_date": stats.get("min_date"),
        "max_date": stats.get("max_date"),
    }


def _payments_list(agent_id: str, start: Optional[datetime], end: Optional[datetime]) -> List[Dict[str, Any]]:
    match: Dict[str, Any] = {
        "agent_id": agent_id,
        "payment_type": {"$ne": "WITHDRAWAL"}
    }
    pipeline: List[Dict[str, Any]] = [
        {"$match": match},
        {"$addFields": {
            "pay_date": {
                "$ifNull": [
                    "$timestamp",
                    {"$dateFromString": {
                        "dateString": "$date",
                        "format": "%Y-%m-%d",
                        "onError": None,
                        "onNull": None
                    }}
                ]
            }
        }},
    ]
    if start and end:
        pipeline.append({"$match": {"pay_date": {"$gte": start, "$lte": end}}})
    pipeline.append({"$sort": {"pay_date": -1}})
    pipeline.append({"$limit": 5000})
    rows = list(payments_col.aggregate(pipeline))
    return rows


def _audits_for_agent(agent_id: str, start: Optional[datetime], end: Optional[datetime]) -> List[Dict[str, Any]]:
    match: Dict[str, Any] = {
        "subject_role": "agent",
        "subject_id": {"$in": [agent_id]}
    }
    if start and end:
        match["incident_date"] = {"$gte": start.strftime("%Y-%m-%d"), "$lte": end.strftime("%Y-%m-%d")}
    docs = list(audits_col.find(match).sort("incident_date", -1))
    return docs


def _agent_id_variants(agent_id: str) -> List[Any]:
    vals: List[Any] = [agent_id]
    try:
        vals.append(ObjectId(agent_id))
    except Exception:
        pass
    return vals


def _customer_counts(agent_id: str) -> Dict[str, int]:
    agent_ids = _agent_id_variants(agent_id)
    active = customers_col.count_documents({"agent_id": {"$in": agent_ids}})
    archived = archived_customers_col.count_documents({"agent_id": {"$in": agent_ids}})
    return {
        "active": int(active),
        "archived": int(archived),
        "total": int(active + archived),
    }


def _agent_customers_rows(agent_id: str, start: Optional[datetime], end: Optional[datetime]) -> List[Dict[str, Any]]:
    agent_ids = _agent_id_variants(agent_id)
    customers = list(customers_col.find(
        {"agent_id": {"$in": agent_ids}},
        {"_id": 1, "name": 1, "phone_number": 1, "location": 1, "purchases": 1}
    ))
    customer_ids = [c["_id"] for c in customers if c.get("_id")]

    pay_map: Dict[str, Dict[str, Any]] = {}
    if customer_ids:
        for batch in _chunks(customer_ids):
            for payment in payments_col.find(
                {
                    "customer_id": {"$in": _customer_id_values(batch)},
                    "$or": [
                        {"payment_type": {"$nin": ["WITHDRAWAL", "SUSU"]}},
                        {"payment_type": "WITHDRAWAL", "product_index": {"$ne": None}},
                    ],
                },
                {"customer_id": 1, "amount": 1, "payment_type": 1, "date": 1, "timestamp": 1},
            ):
                cid = str(payment.get("customer_id") or "")
                pay = pay_map.setdefault(cid, {"total_paid": 0.0, "last_payment": None, "last_amount": 0.0})
                amount = _safe_float(payment.get("amount"), 0.0)
                if payment.get("payment_type") == "WITHDRAWAL":
                    amount *= -1
                pay["total_paid"] = round(_safe_float(pay.get("total_paid"), 0.0) + amount, 2)

                if payment.get("payment_type") not in ("WITHDRAWAL", "SUSU"):
                    pay_date = payment.get("timestamp") or payment.get("date")
                    current_label = _safe_date_str(pay_date)
                    existing_label = _safe_date_str(pay.get("last_payment"))
                    if not pay.get("last_payment") or current_label >= existing_label:
                        pay["last_payment"] = pay_date
                        pay["last_amount"] = _safe_float(payment.get("amount"), 0.0)

    rows: List[Dict[str, Any]] = []
    for c in customers:
        cid = str(c.get("_id"))
        purchases = c.get("purchases") or []
        product_names = []
        for p in purchases:
            name = (p.get("product") or {}).get("name")
            if name:
                product_names.append(name)
        pay = pay_map.get(cid, {"total_paid": 0, "last_payment": None})
        rows.append({
            "customer_name": c.get("name") or "N/A",
            "phone": c.get("phone_number") or "N/A",
            "location": c.get("location") or "N/A",
            "products": ", ".join(product_names) if product_names else "N/A",
            "total_paid": float(pay.get("total_paid") or 0),
            "last_payment": _safe_date_str(pay.get("last_payment")),
            "last_amount": float(pay.get("last_amount") or 0)
        })
    return rows


def _set_job(job_id: str, **updates):
    with JOB_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(updates)


def _build_pdf_to_path(
    file_path: str,
    exec_doc: Dict[str, Any],
    agent: Dict[str, Any],
    manager: Optional[Dict[str, Any]],
    audits: List[Dict[str, Any]],
    pay_stats: Dict[str, Any],
    pay_rows: List[Dict[str, Any]],
    cust_counts: Dict[str, int],
    cust_rows: List[Dict[str, Any]],
    range_label: str,
    statement_id: str,
    progress_cb
):
    progress_cb(10)
    total_paid = pay_stats["total_amount"]
    total_txn = pay_stats["count"]
    min_d = pay_stats["min_date"]
    max_d = pay_stats["max_date"]
    if isinstance(min_d, datetime) and isinstance(max_d, datetime):
        days = max((max_d.date() - min_d.date()).days + 1, 1)
    else:
        days = 1
    avg_daily = total_paid / days if days else 0
    avg_txn = total_paid / total_txn if total_txn else 0

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elems: List[Any] = []

    progress_cb(20)
    logo_img = None
    try:
        logo_resp = requests.get(LOGO_URL, timeout=8)
        logo_resp.raise_for_status()
        logo_img = Image(BytesIO(logo_resp.content), width=1.0 * inch, height=1.0 * inch)
    except Exception:
        pass

    # QR Code with statement details
    qr_payload = f"Company:{COMPANY_NAME}|StatementID:{statement_id}|Date:{datetime.utcnow().strftime('%Y-%m-%d')}|Agent:{agent.get('name') or agent.get('username')}"
    qr_code = qr.QrCodeWidget(qr_payload)
    bounds = qr_code.getBounds()
    size = 80
    d = Drawing(size, size, transform=[size / (bounds[2]-bounds[0]), 0, 0, size / (bounds[3]-bounds[1]), 0, 0])
    d.add(qr_code)

    header_left = []
    if logo_img:
        header_left.append(logo_img)
    header_left.append(Paragraph(f"<b>{COMPANY_NAME}</b>", styles["Heading2"]))
    header_left.append(Paragraph("Agent Audit Statement", styles["Heading3"]))

    header_right = [
        Paragraph(f"<b>Statement ID:</b> {statement_id}", styles["Normal"]),
        Paragraph(f"<b>Statement Date:</b> {datetime.utcnow().strftime('%Y-%m-%d')}", styles["Normal"]),
        Paragraph(f"<b>Period:</b> {range_label}", styles["Normal"]),
    ]

    header_table = Table(
        [[header_left, header_right, d]],
        colWidths=[250, 180, 90]
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems += [header_table, Spacer(1, 12)]

    profile_img = None
    try:
        img_url = (agent.get("image_url") or "").strip()
        if img_url:
            resp = requests.get(img_url, timeout=8)
            resp.raise_for_status()
            profile_img = Image(BytesIO(resp.content), width=0.9 * inch, height=0.9 * inch)
    except Exception:
        profile_img = None

    info_data = [
        ["Agent Name", agent.get("name") or agent.get("username") or "N/A"],
        ["Agent ID", str(agent.get("_id"))],
        ["Role", agent.get("role") or "N/A"],
        ["Branch", agent.get("branch") or "N/A"],
        ["Phone", agent.get("phone") or "N/A"],
        ["Email", agent.get("email") or "N/A"],
        ["Gender", agent.get("gender") or "N/A"],
        ["Position", agent.get("position") or "N/A"],
        ["Location", agent.get("location") or "N/A"],
        ["Start Date", agent.get("start_date") or "N/A"],
        ["Date Registered", agent.get("date_registered") or "N/A"],
        ["Assets", ", ".join(agent.get("assets") or []) or "N/A"],
        ["Profile Image", profile_img if profile_img else "N/A"],
        ["Status", agent.get("status") or "N/A"],
        ["Manager", (manager or {}).get("name") or "N/A"],
        ["Statement Period", range_label],
    ]
    info_table = Table(info_data, colWidths=[160, 340])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems += [info_table, Spacer(1, 12)]

    summary_data = [
        ["Sales Total", f"GHS {total_paid:,.2f}"],
        ["Sales Transactions", str(total_txn)],
        ["Average Daily Sales", f"GHS {avg_daily:,.2f}"],
        ["Average Per Transaction", f"GHS {avg_txn:,.2f}"],
        ["Total Customers", str(cust_counts["total"])],
        ["Active Customers", str(cust_counts["active"])],
        ["Archived Customers", str(cust_counts["archived"])],
    ]
    summary_table = Table(summary_data, colWidths=[200, 300])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems += [Paragraph("Summary", styles["Heading3"]), summary_table, Spacer(1, 12)]

    progress_cb(55)
    audit_rows = [["Date", "Type", "Description", "Lost", "Paid", "Left", "Status"]]
    for d in audits:
        audit_rows.append([
            d.get("incident_date", ""),
            d.get("issue_type", ""),
            d.get("description", ""),
            f"{float(d.get('amount_lost', 0) or 0):,.2f}",
            f"{float(d.get('amount_paid', 0) or 0):,.2f}",
            f"{float(d.get('amount_left', 0) or 0):,.2f}",
            d.get("status", "Open"),
        ])
    if len(audit_rows) == 1:
        audit_rows.append(["-", "-", "No audit records", "0.00", "0.00", "0.00", "-"])

    audit_table = Table(audit_rows, repeatRows=1, colWidths=[70, 70, 220, 55, 55, 55, 55])
    audit_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems += [Paragraph("Audit Findings (Stolen / Lost)", styles["Heading3"]), audit_table, Spacer(1, 12)]

    progress_cb(75)
    pay_table_rows = [["Date", "Customer ID", "Type", "Amount"]]
    for p in pay_rows[:200]:
        pay_table_rows.append([
            _safe_date_str(p.get("pay_date") or p.get("date")),
            str(p.get("customer_id") or ""),
            p.get("payment_type") or "",
            f"{float(p.get('amount', 0) or 0):,.2f}",
        ])
    if len(pay_table_rows) == 1:
        pay_table_rows.append(["-", "-", "-", "0.00"])

    pay_table = Table(pay_table_rows, repeatRows=1, colWidths=[80, 200, 70, 80])
    pay_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elems += [Paragraph("Payments (Sample)", styles["Heading3"]), pay_table]

    progress_cb(90)
    cust_table_rows = [["Customer Name", "Phone", "Location", "Product(s)", "Total Paid", "Last Payment", "Last Amount"]]
    for r in cust_rows:
        cust_table_rows.append([
            r.get("customer_name", ""),
            r.get("phone", ""),
            r.get("location", ""),
            r.get("products", ""),
            f"{float(r.get('total_paid', 0) or 0):,.2f}",
            r.get("last_payment", "") or "-",
            f"{float(r.get('last_amount', 0) or 0):,.2f}"
        ])
    if len(cust_table_rows) == 1:
        cust_table_rows.append(["-", "-", "-", "-", "0.00", "-", "0.00"])

    cust_table = Table(cust_table_rows, repeatRows=1, colWidths=[95, 85, 85, 130, 60, 60, 60])
    cust_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elems += [Paragraph("Agent Customers", styles["Heading3"]), cust_table]

    progress_cb(95)
    doc.build(elems)
    with open(file_path, "wb") as f:
        f.write(buf.getvalue())
    progress_cb(100)


def _build_excel_to_path(
    file_path: str,
    agent: Dict[str, Any],
    manager: Optional[Dict[str, Any]],
    audits: List[Dict[str, Any]],
    pay_stats: Dict[str, Any],
    pay_rows: List[Dict[str, Any]],
    cust_counts: Dict[str, int],
    cust_rows: List[Dict[str, Any]],
    customer_card_tabs: Dict[str, List[Dict[str, Any]]],
    range_label: str,
    progress_cb
):
    if openpyxl is None:
        raise RuntimeError("openpyxl missing")

    progress_cb(15)
    total_paid = pay_stats["total_amount"]
    total_txn = pay_stats["count"]
    min_d = pay_stats["min_date"]
    max_d = pay_stats["max_date"]
    if isinstance(min_d, datetime) and isinstance(max_d, datetime):
        days = max((max_d.date() - min_d.date()).days + 1, 1)
    else:
        days = 1
    avg_daily = total_paid / days if days else 0
    avg_txn = total_paid / total_txn if total_txn else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="E2E8F0")

    ws["A1"] = "Agent Audit Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Agent Name"
    ws["B3"] = agent.get("name") or agent.get("username") or "N/A"
    ws["A4"] = "Agent ID"
    ws["B4"] = str(agent.get("_id"))
    ws["A5"] = "Role"
    ws["B5"] = agent.get("role") or "N/A"
    ws["A6"] = "Branch"
    ws["B6"] = agent.get("branch") or "N/A"
    ws["A7"] = "Phone"
    ws["B7"] = agent.get("phone") or "N/A"
    ws["A8"] = "Email"
    ws["B8"] = agent.get("email") or "N/A"
    ws["A9"] = "Gender"
    ws["B9"] = agent.get("gender") or "N/A"
    ws["A10"] = "Position"
    ws["B10"] = agent.get("position") or "N/A"
    ws["A11"] = "Location"
    ws["B11"] = agent.get("location") or "N/A"
    ws["A12"] = "Start Date"
    ws["B12"] = agent.get("start_date") or "N/A"
    ws["A13"] = "Date Registered"
    ws["B13"] = agent.get("date_registered") or "N/A"
    ws["A14"] = "Assets"
    ws["B14"] = ", ".join(agent.get("assets") or []) or "N/A"
    ws["A15"] = "Profile Image"
    ws["B15"] = agent.get("image_url") or "N/A"
    ws["A16"] = "Manager"
    ws["B16"] = (manager or {}).get("name") or "N/A"
    ws["A17"] = "Statement Period"
    ws["B17"] = range_label

    ws["A19"] = "Sales Total"
    ws["B19"] = total_paid
    ws["A20"] = "Sales Transactions"
    ws["B20"] = total_txn
    ws["A21"] = "Average Daily Sales"
    ws["B21"] = avg_daily
    ws["A22"] = "Average Per Transaction"
    ws["B22"] = avg_txn
    ws["A23"] = "Total Customers"
    ws["B23"] = cust_counts["total"]
    ws["A24"] = "Active Customers"
    ws["B24"] = cust_counts["active"]
    ws["A25"] = "Archived Customers"
    ws["B25"] = cust_counts["archived"]
    ws["A27"] = "Total Product Cards"
    ws["B27"] = sum(len(customer_card_tabs.get(key, [])) for key in ("ongoing", "closed", "completed"))
    ws["A28"] = "Ongoing Product Cards"
    ws["B28"] = len(customer_card_tabs.get("ongoing", []))
    ws["A29"] = "Closed Product Cards"
    ws["B29"] = len(customer_card_tabs.get("closed", []))
    ws["A30"] = "Completed Product Cards"
    ws["B30"] = len(customer_card_tabs.get("completed", []))

    ws.sheet_properties.tabColor = "2563EB"
    for row in range(3, 31):
        ws[f"A{row}"].font = header_font
    for row in range(27, 31):
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="DBEAFE")
        ws[f"B{row}"].fill = PatternFill("solid", fgColor="EFF6FF")

    progress_cb(45)
    ws_a = wb.create_sheet("Audit Findings")
    ws_a.sheet_properties.tabColor = "F97316"
    ws_a.append(["Date", "Type", "Description", "Lost", "Paid", "Left", "Status"])
    for cell in ws_a[1]:
        cell.font = header_font
        cell.fill = fill
    for d in audits:
        ws_a.append([
            d.get("incident_date", ""),
            d.get("issue_type", ""),
            d.get("description", ""),
            float(d.get("amount_lost", 0) or 0),
            float(d.get("amount_paid", 0) or 0),
            float(d.get("amount_left", 0) or 0),
            d.get("status", "Open"),
        ])

    progress_cb(70)
    ws_p = wb.create_sheet("Payments")
    ws_p.sheet_properties.tabColor = "7C3AED"
    ws_p.append(["Date", "Customer ID", "Type", "Amount"])
    for cell in ws_p[1]:
        cell.font = header_font
        cell.fill = fill
    for p in pay_rows:
        ws_p.append([
            _safe_date_str(p.get("pay_date") or p.get("date")),
            str(p.get("customer_id") or ""),
            p.get("payment_type") or "",
            float(p.get("amount", 0) or 0),
        ])

    customer_headers = [
        "Customer Name",
        "Phone",
        "Location",
        "Branch",
        "Manager",
        "Product",
        "Product #",
        "Product Status",
        "Package Status",
        "Total",
        "Paid",
        "Left",
        "Purchase Date",
    ]

    def _add_customer_state_sheet(sheet_name: str, rows: List[Dict[str, Any]], tab_color: str, header_color: str):
        ws_state = wb.create_sheet(sheet_name)
        ws_state.sheet_properties.tabColor = tab_color
        ws_state.append(customer_headers)
        state_fill = PatternFill("solid", fgColor=header_color)
        for cell in ws_state[1]:
            cell.font = header_font
            cell.fill = state_fill
        for row in rows:
            ws_state.append([
                row.get("customer_name", ""),
                row.get("phone", ""),
                row.get("location", ""),
                row.get("branch", ""),
                row.get("manager_name", ""),
                row.get("product_name", ""),
                int(row.get("product_index", 0) or 0) + 1,
                row.get("product_status", ""),
                row.get("package_status", ""),
                float(row.get("total", 0) or 0),
                float(row.get("paid", 0) or 0),
                float(row.get("left", 0) or 0),
                row.get("purchase_date", ""),
            ])
        if not rows:
            ws_state.append(["No records", "", "", "", "", "", "", "", "", 0, 0, 0, ""])

    _add_customer_state_sheet("Ongoing Customers", customer_card_tabs.get("ongoing", []), "0EA5E9", "E0F2FE")
    _add_customer_state_sheet("Closed Customers", customer_card_tabs.get("closed", []), "DC2626", "FEE2E2")
    _add_customer_state_sheet("Completed Customers", customer_card_tabs.get("completed", []), "16A34A", "DCFCE7")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    progress_cb(90)
    wb.save(file_path)
    progress_cb(100)

@executive_auditing_bp.route("", methods=["GET"])
@executive_auditing_bp.route("/", methods=["GET"])
def auditing_page():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope
    _, exec_doc = scope

    # Users grouped by role for the form
    users_by_role = {"manager": [], "admin": [], "agent": []}
    for role in users_by_role.keys():
        cursor = users_col.find(
            {"role": {"$regex": f"^{role}$", "$options": "i"}},
            {"_id": 1, "name": 1, "username": 1, "branch": 1}
        ).sort("name", 1)
        for u in cursor:
            users_by_role[role].append({
                "id": str(u["_id"]),
                "label": _subject_label(u),
                "branch": u.get("branch", "")
            })

    match = _build_match(request.args)
    docs = list(audits_col.find(match).sort("created_at", -1).limit(300))

    rows = []
    for d in docs:
        rows.append({
            "_id": str(d.get("_id")),
            "incident_date": d.get("incident_date", ""),
            "subject_role": (d.get("subject_role") or "").title(),
            "subject_name": d.get("subject_name", ""),
            "subject_branch": d.get("subject_branch", ""),
            "issue_type": d.get("issue_type", ""),
            "description": d.get("description", ""),
            "amount_lost": _safe_float(d.get("amount_lost", 0)),
            "amount_paid": _safe_float(d.get("amount_paid", 0)),
            "amount_left": _safe_float(d.get("amount_left", 0)),
            "status": d.get("status", "Open"),
        })

    # Totals for current filter
    pipeline = []
    if match:
        pipeline.append({"$match": match})
    pipeline.append({
        "$group": {
            "_id": None,
            "total_lost": {"$sum": {"$ifNull": ["$amount_lost", 0]}},
            "total_paid": {"$sum": {"$ifNull": ["$amount_paid", 0]}},
            "total_left": {"$sum": {"$ifNull": ["$amount_left", 0]}},
            "count": {"$sum": 1}
        }
    })
    totals_agg = list(audits_col.aggregate(pipeline))
    totals = totals_agg[0] if totals_agg else {"total_lost": 0, "total_paid": 0, "total_left": 0, "count": 0}

    # Top 3 owing users this month
    now = datetime.utcnow()
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_str = start_month.strftime("%Y-%m-%d")
    top_match = {"amount_left": {"$gt": 0}}
    top_pipeline = [
        {"$match": {
            "incident_date": {"$gte": month_str, "$lte": now.strftime("%Y-%m-%d")},
            "amount_left": {"$gt": 0}
        }},
        {"$group": {
            "_id": "$subject_id",
            "subject_name": {"$first": "$subject_name"},
            "subject_role": {"$first": "$subject_role"},
            "total_left": {"$sum": {"$ifNull": ["$amount_left", 0]}}
        }},
        {"$sort": {"total_left": -1}},
        {"$limit": 3}
    ]
    top_owing = [
        {
            "name": t.get("subject_name") or "User",
            "role": (t.get("subject_role") or "").title(),
            "amount": float(t.get("total_left") or 0)
        }
        for t in audits_col.aggregate(top_pipeline)
    ]

    # Amount lost this month
    month_total = list(audits_col.aggregate([
        {"$match": {"incident_date": {"$gte": month_str, "$lte": now.strftime("%Y-%m-%d")}}},
        {"$group": {"_id": None, "lost": {"$sum": {"$ifNull": ["$amount_lost", 0]}}}}
    ]))
    lost_this_month = float(month_total[0]["lost"]) if month_total else 0.0

    return render_template(
        "executive/auditing.html",
        executive_name=exec_doc.get("name", "Executive"),
        users_by_role=users_by_role,
        rows=rows,
        totals=totals,
        top_owing=top_owing,
        lost_this_month=lost_this_month,
        today=datetime.utcnow().strftime("%Y-%m-%d"),
        current_role=(request.args.get("role") or ""),
        current_status=(request.args.get("status") or "")
    )


@executive_auditing_bp.route("/new", methods=["POST"])
def auditing_new():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope
    exec_id, _ = scope

    subject_role = (request.form.get("subject_role") or "").strip().lower()
    subject_id = (request.form.get("subject_id") or "").strip()
    issue_type = (request.form.get("issue_type") or "").strip()
    description = (request.form.get("description") or "").strip()
    incident_date = (request.form.get("incident_date") or "").strip()
    amount_lost = _safe_float(request.form.get("amount_lost"), 0)
    amount_paid = _safe_float(request.form.get("amount_paid"), 0)

    if subject_role not in ("manager", "admin", "agent") or not subject_id or amount_lost <= 0:
        flash("Provide role, user, and a valid amount.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    try:
        subject_user = users_col.find_one({"_id": ObjectId(subject_id)})
    except Exception:
        subject_user = users_col.find_one({"_id": subject_id})

    subject_name = _subject_label(subject_user or {})
    subject_branch = (subject_user or {}).get("branch", "")

    amount_left = max(amount_lost - amount_paid, 0)
    if amount_left <= 0:
        status = "Settled"
    elif amount_paid > 0:
        status = "Partial"
    else:
        status = "Open"

    audits_col.insert_one({
        "subject_role": subject_role,
        "subject_id": subject_id,
        "subject_name": subject_name,
        "subject_branch": subject_branch,
        "issue_type": issue_type,
        "description": description,
        "incident_date": incident_date,
        "amount_lost": amount_lost,
        "amount_paid": amount_paid,
        "amount_left": amount_left,
        "status": status,
        "recorded_by": exec_id,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    })

    flash("Audit record saved.", "success")
    return redirect(url_for("executive_auditing.auditing_page"))


@executive_auditing_bp.route("/update_paid", methods=["POST"])
def auditing_update_paid():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope

    audit_id = (request.form.get("audit_id") or "").strip()
    update_field = (request.form.get("update_field") or "paid").strip().lower()
    if update_field not in ("lost", "paid"):
        update_field = "paid"

    if "amount_value" in request.form:
        amount_value = _safe_float(request.form.get("amount_value"), 0)
    elif update_field == "lost":
        amount_value = _safe_float(request.form.get("amount_lost"), 0)
    else:
        amount_value = _safe_float(request.form.get("amount_paid"), 0)

    if not audit_id:
        flash("Missing record id.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))
    if amount_value < 0:
        flash("Amount cannot be negative.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    try:
        doc = audits_col.find_one({"_id": ObjectId(audit_id)})
    except Exception:
        doc = None

    if not doc:
        flash("Record not found.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    amount_lost = _safe_float(doc.get("amount_lost", 0))
    amount_paid = _safe_float(doc.get("amount_paid", 0))
    if update_field == "lost":
        amount_lost = amount_value
    else:
        amount_paid = amount_value

    amount_left = max(amount_lost - amount_paid, 0)
    if amount_left <= 0:
        status = "Settled"
    elif amount_paid > 0:
        status = "Partial"
    else:
        status = "Open"

    audits_col.update_one(
        {"_id": ObjectId(audit_id)},
        {"$set": {
            "amount_lost": amount_lost,
            "amount_paid": amount_paid,
            "amount_left": amount_left,
            "status": status,
            "updated_at": datetime.utcnow()
        }}
    )

    flash("Audit record updated.", "success")
    return redirect(url_for("executive_auditing.auditing_page"))


@executive_auditing_bp.route("/delete", methods=["POST"])
def auditing_delete():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope

    audit_id = (request.form.get("audit_id") or "").strip()
    if not audit_id:
        flash("Missing record id.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    try:
        audits_col.delete_one({"_id": ObjectId(audit_id)})
    except Exception:
        flash("Invalid record id.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    flash("Record deleted.", "success")
    return redirect(url_for("executive_auditing.auditing_page"))


@executive_auditing_bp.route("/statement/pdf", methods=["GET"])
def auditing_statement_pdf():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope
    _, exec_doc = scope

    agent_id = (request.args.get("agent_id") or "").strip()
    range_key = request.args.get("range", "all_time")
    start_str = request.args.get("start") or ""
    end_str = request.args.get("end") or ""

    agent = _agent_doc(agent_id)
    if not agent:
        flash("Agent not found.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    start, end, range_label = _parse_range(range_key, start_str, end_str)
    manager = _manager_doc(agent.get("manager_id"))
    audits = _audits_for_agent(agent_id, start, end)
    pay_stats = _payments_stats(agent_id, start, end)
    pay_rows = _payments_list(agent_id, start, end)
    cust_counts = _customer_counts(agent_id)
    cust_rows = _agent_customers_rows(agent_id, start, end)

    filename = f"agent_audit_statement_{agent_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    tmp_path = os.path.join(EXPORT_DIR, filename)
    statement_id = f"AUD-{uuid.uuid4().hex[:8].upper()}"
    _build_pdf_to_path(
        tmp_path, exec_doc, agent, manager, audits, pay_stats, pay_rows, cust_counts, cust_rows, range_label,
        statement_id,
        lambda p: None
    )
    return send_file(tmp_path, as_attachment=True, download_name=filename, mimetype="application/pdf")


@executive_auditing_bp.route("/statement/excel", methods=["GET"])
def auditing_statement_excel():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return scope

    if openpyxl is None:
        flash("Excel export is not available. Missing openpyxl.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    agent_id = (request.args.get("agent_id") or "").strip()
    range_key = request.args.get("range", "all_time")
    start_str = request.args.get("start") or ""
    end_str = request.args.get("end") or ""

    agent = _agent_doc(agent_id)
    if not agent:
        flash("Agent not found.", "danger")
        return redirect(url_for("executive_auditing.auditing_page"))

    start, end, range_label = _parse_range(range_key, start_str, end_str)
    manager = _manager_doc(agent.get("manager_id"))
    audits = _audits_for_agent(agent_id, start, end)
    pay_stats = _payments_stats(agent_id, start, end)
    pay_rows = _payments_list(agent_id, start, end)
    cust_counts = _customer_counts(agent_id)
    cust_rows = _agent_customers_rows(agent_id, start, end)
    customer_card_tabs, _ = _build_customer_card_tabs(agent_id=agent_id)

    filename = f"agent_audit_statement_{agent_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    tmp_path = os.path.join(EXPORT_DIR, filename)
    _build_excel_to_path(
        tmp_path, agent, manager, audits, pay_stats, pay_rows, cust_counts, cust_rows, customer_card_tabs, range_label,
        lambda p: None
    )
    return send_file(tmp_path, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@executive_auditing_bp.route("/statement/start", methods=["POST"])
def auditing_statement_start():
    scope = _ensure_exec_or_redirect()
    if not isinstance(scope, tuple):
        return jsonify(ok=False, message="Unauthorized"), 401
    exec_id, exec_doc = scope

    fmt = (request.form.get("format") or "").lower()
    agent_id = (request.form.get("agent_id") or "").strip()
    range_key = request.form.get("range", "all_time")
    start_str = request.form.get("start") or ""
    end_str = request.form.get("end") or ""

    if fmt not in ("pdf", "excel"):
        return jsonify(ok=False, message="Invalid format"), 400

    agent = _agent_doc(agent_id)
    if not agent:
        return jsonify(ok=False, message="Agent not found"), 404

    if fmt == "excel" and openpyxl is None:
        return jsonify(ok=False, message="Excel export unavailable"), 400

    job_id = uuid.uuid4().hex
    filename = f"agent_audit_statement_{agent_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{ 'pdf' if fmt=='pdf' else 'xlsx' }"
    file_path = os.path.join(EXPORT_DIR, filename)

    with JOB_LOCK:
        JOBS[job_id] = {
            "status": "running",
            "progress": 0,
            "file_path": file_path,
            "filename": filename,
            "format": fmt,
            "created_at": datetime.utcnow(),
        }

    def worker():
        try:
            start, end, range_label = _parse_range(range_key, start_str, end_str)
            manager = _manager_doc(agent.get("manager_id"))
            audits = _audits_for_agent(agent_id, start, end)
            pay_stats = _payments_stats(agent_id, start, end)
            pay_rows = _payments_list(agent_id, start, end)
            cust_counts = _customer_counts(agent_id)
            cust_rows = _agent_customers_rows(agent_id, start, end)

            def progress_cb(p):
                _set_job(job_id, progress=p)

            if fmt == "pdf":
                statement_id = f"AUD-{job_id[:8].upper()}"
                _build_pdf_to_path(file_path, exec_doc, agent, manager, audits, pay_stats, pay_rows, cust_counts, cust_rows, range_label, statement_id, progress_cb)
            else:
                customer_card_tabs, _ = _build_customer_card_tabs(agent_id=agent_id)
                _build_excel_to_path(file_path, agent, manager, audits, pay_stats, pay_rows, cust_counts, cust_rows, customer_card_tabs, range_label, progress_cb)
            _set_job(job_id, status="done", progress=100)
        except Exception as e:
            _set_job(job_id, status="error", message=str(e))

    threading.Thread(target=worker, daemon=True).start()
    return jsonify(ok=True, job_id=job_id)


@executive_auditing_bp.route("/statement/progress/<job_id>", methods=["GET"])
def auditing_statement_progress(job_id):
    with JOB_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify(ok=False, message="Job not found"), 404
    return jsonify(ok=True, status=job.get("status"), progress=job.get("progress", 0), message=job.get("message"))


@executive_auditing_bp.route("/statement/download/<job_id>", methods=["GET"])
def auditing_statement_download(job_id):
    with JOB_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify(ok=False, message="Job not found"), 404
    if job.get("status") != "done":
        return jsonify(ok=False, message="File not ready"), 400
    return send_file(job["file_path"], as_attachment=True, download_name=job["filename"])
